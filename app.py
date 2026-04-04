import os
import io
import base64
import calendar
import requests
import pandas as pd
from datetime import datetime, timedelta, date
from flask import Flask, request, jsonify, render_template_string

app = Flask(__name__)

RESEND_API_KEY = os.environ.get("RESEND_API_KEY", "")
FROM_EMAIL     = os.environ.get("FROM_EMAIL", "")
DEFAULT_TO     = os.environ.get("DEFAULT_TO", "")
DEFAULT_CC     = os.environ.get("DEFAULT_CC", "")

# ── date helpers ──────────────────────────────────────────────────────────────

def parse_date(raw):
    """Parse to date-only (no time). Returns datetime.date or None."""
    if pd.isna(raw):
        return None
    s = str(raw).strip()
    if s in ("--", "", "nan", "None"):
        return None
    for fmt in ("%d-%m-%Y %H:%M", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    for fmt in ("%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            pass
    return None

def fmt_date(d):
    if d is None:
        return ""
    if isinstance(d, datetime):
        return d.strftime("%d-%m-%Y 00:00")
    return d.strftime("%d-%m-%Y") + " 00:00"

def parse_anchor(s):
    """Parse user-entered anchor (dd-mm-yyyy). Returns datetime.date."""
    s = s.strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    raise ValueError(f"Cannot parse anchor date: '{s}'. Use dd-mm-yyyy format.")

def get_upcoming_window(anchor):
    """
    anchor.day == 1  → window: 1st to 15th of anchor month
    anchor.day == 16 → window: 16th to last day of anchor month
    Returns (start, end) as date objects.
    """
    if anchor.day == 1:
        return anchor, anchor.replace(day=15)
    else:
        last_day = calendar.monthrange(anchor.year, anchor.month)[1]
        return anchor, anchor.replace(day=last_day)

# ── Excel loader ──────────────────────────────────────────────────────────────

def find_header_row(raw_df):
    for i, row in raw_df.iterrows():
        if any(str(v).strip() == "Installation Date" for v in row.values):
            return i
    return None

def load_dataframe(file_obj):
    raw = pd.read_excel(file_obj, header=None)
    header_row = find_header_row(raw)

    if header_row is None:
        # Try as simple 3-col file (no header)
        file_obj.seek(0)
        df = pd.read_excel(file_obj, header=None)
        if df.shape[1] >= 3:
            out = pd.DataFrame()
            out["company"] = df.iloc[:, 0].astype(str).str.strip()
            out["plate"]   = df.iloc[:, 1].astype(str).str.strip()
            out["date"]    = df.iloc[:, 2].apply(parse_date)
            out = out[out["company"].notna() & (out["company"] != "nan") & (out["company"] != "")]
            out = out.dropna(subset=["date"])
            return out
        raise ValueError("Could not find 'Installation Date' column in this file.")

    file_obj.seek(0)
    df = pd.read_excel(file_obj, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    col_map = {c.lower(): c for c in df.columns}

    def get_col(name):
        key = name.lower()
        if key not in col_map:
            raise ValueError(f"Column '{name}' not found. Available: {list(df.columns)}")
        return col_map[key]

    company_col = get_col("Company")
    plate_col   = get_col("Plate Number")
    date_col    = get_col("Installation Date")

    out = pd.DataFrame()
    out["company"] = df[company_col].astype(str).str.strip()
    out["plate"]   = df[plate_col].astype(str).str.strip()
    out["date"]    = df[date_col].apply(parse_date)
    out = out[out["company"].notna() & (out["company"] != "nan") & (out["company"] != "")]
    out = out.dropna(subset=["date"])
    return out

# ── bucketing ─────────────────────────────────────────────────────────────────

def bucket_records(df, anchor):
    """
    anchor: datetime.date, must be 1st or 16th of month.

    Upcoming  : anchor → upcoming_end  (forward half-month window)
    Band 1 (last 30d)  : anchor-30 to anchor-1   (most recent expired)
    Band 2 (last 60d)  : anchor-60 to anchor-31
    Band 3 (last 90d)  : anchor-90 to anchor-61
    Band 4 (last 120d) : anchor-120 to anchor-91
    Band 5 (last 150d) : anchor-150 to anchor-121

    All comparisons are date-only (no time).
    """
    up_start, up_end = get_upcoming_window(anchor)

    b1_end   = anchor - timedelta(days=1)
    b1_start = anchor - timedelta(days=30)
    b2_end   = anchor - timedelta(days=31)
    b2_start = anchor - timedelta(days=60)
    b3_end   = anchor - timedelta(days=61)
    b3_start = anchor - timedelta(days=90)
    b4_end   = anchor - timedelta(days=91)
    b4_start = anchor - timedelta(days=120)
    b5_end   = anchor - timedelta(days=121)
    b5_start = anchor - timedelta(days=150)

    buckets = {
        "upcoming": [],
        "b1": [], "b2": [], "b3": [], "b4": [], "b5": [],
        "_anchor":   anchor,
        "_up_start": up_start,
        "_up_end":   up_end,
        "_b1_start": b1_start, "_b1_end": b1_end,
        "_b2_start": b2_start, "_b2_end": b2_end,
        "_b3_start": b3_start, "_b3_end": b3_end,
        "_b4_start": b4_start, "_b4_end": b4_end,
        "_b5_start": b5_start, "_b5_end": b5_end,
    }

    for _, row in df.iterrows():
        d = row["date"]
        if d is None:
            continue
        if up_start <= d <= up_end:
            buckets["upcoming"].append(row)
        elif b1_start <= d <= b1_end:
            buckets["b1"].append(row)
        elif b2_start <= d <= b2_end:
            buckets["b2"].append(row)
        elif b3_start <= d <= b3_end:
            buckets["b3"].append(row)
        elif b4_start <= d <= b4_end:
            buckets["b4"].append(row)
        elif b5_start <= d <= b5_end:
            buckets["b5"].append(row)

    return buckets

# ── report builders ───────────────────────────────────────────────────────────

def company_summary(records):
    counts = {}
    for r in records:
        counts[r["company"]] = counts.get(r["company"], 0) + 1
    items = sorted(counts.items())
    items.append(("Total", sum(c for _, c in items)))
    return items

def build_table_text(headers, rows, col_widths):
    lines = []
    header_line = "  ".join(h.ljust(w) for h, w in zip(headers, col_widths))
    lines.append(header_line)
    lines.append("-" * len(header_line))
    for row in rows:
        lines.append("  ".join(str(v).ljust(w) for v, w in zip(row, col_widths)))
    return "\n".join(lines)

def date_range_label(start, end):
    return f"{start.strftime('%d-%m-%Y')} to {end.strftime('%d-%m-%Y')}"

def build_email_body(buckets, report_date_str):
    anchor   = buckets["_anchor"]
    up_start = buckets["_up_start"]
    up_end   = buckets["_up_end"]

    parts = []
    parts.append(f"Device Expiry Report - {report_date_str}")
    parts.append(f"Anchor: {anchor.strftime('%d-%m-%Y')}")
    parts.append("=" * 62)

    # ── Upcoming ──
    up = buckets["upcoming"]
    up_label = f"Upcoming Device Expiry ({date_range_label(up_start, up_end)})"
    parts.append(f"\n{up_label}\n")
    parts.append(build_table_text(["Company", "Device Count"], company_summary(up), [32, 12]))
    parts.append("")
    parts.append(build_table_text(
        ["Company", "Plate Number", "Installation Date"],
        [(r["company"], r["plate"], fmt_date(r["date"])) for r in up], [32, 18, 20]
    ))

    # ── Expired bands ──
    parts.append("\n\nExpired Devices (Grouped by Months)")
    parts.append("=" * 62)

    expired_bands = [
        ("b1", "Expired within last 30 days"),
        ("b2", "Expired within last 60 days"),
        ("b3", "Expired within last 90 days"),
        ("b4", "Expired within last 120 days"),
        ("b5", "Expired within last 150 days"),
    ]

    for key, label in expired_bands:
        recs = buckets[key]
        if not recs:
            continue
        start = buckets[f"_{key}_start"]
        end   = buckets[f"_{key}_end"]
        parts.append(f"\n{label} ({date_range_label(start, end)})\n")
        parts.append(build_table_text(["Company", "Device Count"], company_summary(recs), [32, 12]))
        parts.append("")
        parts.append(build_table_text(
            ["Company", "Plate Number", "Installation Date"],
            [(r["company"], r["plate"], fmt_date(r["date"])) for r in recs], [32, 18, 20]
        ))

    return "\n".join(parts)

def build_html_body(buckets, report_date_str):
    anchor   = buckets["_anchor"]
    up_start = buckets["_up_start"]
    up_end   = buckets["_up_end"]

    def th(bg="#d6e4f0"):
        return f"style='background:{bg}'"

    def summary_table(records):
        rows = "".join(
            f"<tr><td>{c}</td><td style='text-align:center'><b>{n}</b></td></tr>"
            for c, n in company_summary(records)
        )
        return f"""<table border='1' cellpadding='5' cellspacing='0'
            style='border-collapse:collapse;width:100%;max-width:520px;margin-bottom:8px'>
          <tr {th()}><th>Company</th><th>Device Count</th></tr>{rows}</table>"""

    def detail_table(records, row_bg="#fce4e4"):
        rows = "".join(
            f"<tr style='background:{row_bg}'><td>{r['company']}</td><td>{r['plate']}</td>"
            f"<td>{fmt_date(r['date'])}</td></tr>"
            for r in records
        )
        return f"""<table border='1' cellpadding='5' cellspacing='0'
            style='border-collapse:collapse;width:100%;margin-bottom:16px'>
          <tr {th()}><th>Company</th><th>Plate Number</th><th>Installation Date</th></tr>
          {rows}
          <tr><td colspan='3'><b>Total: {len(records)}</b></td></tr></table>"""

    expired_bands = [
        ("b1", "Expired within last 30 days"),
        ("b2", "Expired within last 60 days"),
        ("b3", "Expired within last 90 days"),
        ("b4", "Expired within last 120 days"),
        ("b5", "Expired within last 150 days"),
    ]

    exp_html = ""
    for key, label in expired_bands:
        recs = buckets[key]
        if not recs:
            continue
        start = buckets[f"_{key}_start"]
        end   = buckets[f"_{key}_end"]
        exp_html += (f"<h3>{label}</h3>"
                     f"<p style='color:#666;font-size:12px'>{date_range_label(start, end)}</p>"
                     f"{summary_table(recs)}{detail_table(recs)}")

    up = buckets["upcoming"]
    return f"""<html><body style='font-family:Arial,sans-serif;font-size:13px;color:#222'>
      <h2>Device Expiry Report - {report_date_str}</h2>
      <p style='color:#555;font-size:12px'>Anchor: {anchor.strftime('%d-%m-%Y')}</p>
      <h3>Upcoming Device Expiry ({date_range_label(up_start, up_end)})</h3>
      {summary_table(up)}{detail_table(up, '#fce4e4')}
      <h2>Expired Devices (Grouped by Days)</h2>
      {exp_html}
    </body></html>"""

# ── result Excel builder ──────────────────────────────────────────────────────

def build_result_excel(buckets, report_date_str):
    """
    Build a formatted Excel workbook mirroring the report structure.
    Returns bytes of the .xlsx file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Expiry Report"

    anchor   = buckets["_anchor"]
    up_start = buckets["_up_start"]
    up_end   = buckets["_up_end"]

    # ── style helpers ──
    BLUE_FILL   = PatternFill("solid", fgColor="D6E4F0")   # header rows
    PINK_FILL   = PatternFill("solid", fgColor="FCE4E4")   # expired detail rows
    GREEN_FILL  = PatternFill("solid", fgColor="D1FAE5")   # upcoming detail rows
    GREY_FILL   = PatternFill("solid", fgColor="F1F5F9")   # total rows
    SECTION_FILL= PatternFill("solid", fgColor="1A1A2E")   # section title

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    bold   = Font(bold=True)
    white  = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    def set_col_widths(widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    set_col_widths([35, 20, 22])

    row = 1

    def write_section_title(title, fill_color="1A1A2E", font_color="FFFFFF"):
        nonlocal row
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws[f"A{row}"]
        cell.value = title
        cell.font = Font(bold=True, color=font_color, size=12)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = center
        row += 1

    def write_report_header():
        nonlocal row
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws[f"A{row}"]
        cell.value = f"Device Expiry Report — {report_date_str}   |   Anchor: {anchor.strftime('%d-%m-%Y')}"
        cell.font = Font(bold=True, size=13, color="1A1A2E")
        cell.alignment = center
        row += 2

    def write_summary_table(records, label):
        nonlocal row
        # Section label
        ws.merge_cells(f"A{row}:C{row}")
        c = ws[f"A{row}"]
        c.value = label
        c.font = Font(bold=True, size=11)
        c.alignment = Alignment(horizontal="left")
        row += 1

        # Summary header
        for col, val in enumerate(["Company", "Device Count"], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = bold; c.fill = BLUE_FILL; c.border = border; c.alignment = center
        row += 1

        summary = company_summary(records)
        for company, count in summary:
            is_total = (company == "Total")
            c1 = ws.cell(row=row, column=1, value=company)
            c2 = ws.cell(row=row, column=2, value=count)
            for c in [c1, c2]:
                c.border = border
                if is_total:
                    c.font = bold; c.fill = GREY_FILL
            row += 1
        row += 1  # blank row

    def write_detail_table(records, detail_fill):
        nonlocal row
        # Detail header
        for col, val in enumerate(["Company", "Plate Number", "Installation Date"], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = bold; c.fill = BLUE_FILL; c.border = border; c.alignment = center
        row += 1

        for r in records:
            vals = [r["company"], r["plate"], fmt_date(r["date"])]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = detail_fill; c.border = border
            row += 1

        # Total row
        ws.merge_cells(f"A{row}:B{row}")
        c1 = ws[f"A{row}"]
        c1.value = "Total"; c1.font = bold; c1.fill = GREY_FILL; c1.border = border
        c3 = ws.cell(row=row, column=3, value=len(records))
        c3.font = bold; c3.fill = GREY_FILL; c3.border = border; c3.alignment = center
        row += 2  # blank row after section

    # ── Write report ──
    write_report_header()

    # Upcoming section
    up = buckets["upcoming"]
    up_label = f"Upcoming Device Expiry  ({date_range_label(up_start, up_end)})"
    write_section_title(up_label)
    write_summary_table(up, "Summary by Company")
    write_detail_table(up, GREEN_FILL)

    # Expired bands
    write_section_title("Expired Devices (Grouped by Days)", fill_color="7F1D1D")
    row += 1

    expired_bands = [
        ("b1", "Expired — Last 30 days"),
        ("b2", "Expired — Last 60 days"),
        ("b3", "Expired — Last 90 days"),
        ("b4", "Expired — Last 120 days"),
        ("b5", "Expired — Last 150 days"),
    ]

    for key, label in expired_bands:
        recs = buckets[key]
        if not recs:
            continue
        start = buckets[f"_{key}_start"]
        end   = buckets[f"_{key}_end"]
        full_label = f"{label}  ({date_range_label(start, end)})"
        write_section_title(full_label, fill_color="4B5563", font_color="FFFFFF")
        write_summary_table(recs, "Summary by Company")
        write_detail_table(recs, PINK_FILL)

    # Freeze header rows
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ── email sender ──────────────────────────────────────────────────────────────

def send_email(to, cc, subject, plain_body, html_body, attachments):
    """
    attachments: list of {"filename": str, "content": bytes}
    """
    payload = {
        "from": FROM_EMAIL, "to": [to],
        "subject": subject, "text": plain_body, "html": html_body,
        "attachments": [
            {"filename": a["filename"],
             "content": base64.b64encode(a["content"]).decode("utf-8")}
            for a in attachments
        ],
    }
    if cc:
        payload["cc"] = [cc]
    resp = requests.post("https://api.resend.com/emails",
        headers={"Authorization": f"Bearer {RESEND_API_KEY}", "Content-Type": "application/json"},
        json=payload, timeout=30)
    if resp.status_code not in (200, 201):
        raise Exception(f"Resend API error {resp.status_code}: {resp.text}")

# ── HTML page ─────────────────────────────────────────────────────────────────

HTML_PAGE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Device Expiry Report</title>
<style>
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: Arial, sans-serif; font-size: 14px; background: #f4f6f9; color: #222; min-height: 100vh; }
  .topbar { background: #1a1a2e; color: #fff; padding: 14px 28px; display: flex; align-items: center; gap: 12px; }
  .topbar h1 { font-size: 17px; font-weight: 600; }
  .topbar .dot { width: 10px; height: 10px; border-radius: 50%; background: #4ade80; }
  .layout { max-width: 860px; margin: 0 auto; padding: 28px 20px; display: grid; gap: 18px; }
  .card { background: #fff; border-radius: 10px; border: 1px solid #e2e8f0; padding: 22px 24px; }
  .card-title { font-size: 13px; font-weight: 600; text-transform: uppercase; letter-spacing: 0.06em; color: #64748b; margin-bottom: 16px; }
  .upload-zone { border: 2px dashed #cbd5e1; border-radius: 8px; padding: 32px 20px; text-align: center; cursor: pointer; transition: border-color 0.2s, background 0.2s; }
  .upload-zone:hover, .upload-zone.drag { border-color: #3b82f6; background: #eff6ff; }
  .upload-zone input { display: none; }
  .upload-icon { font-size: 36px; margin-bottom: 8px; }
  .upload-zone p { color: #475569; font-size: 14px; }
  .upload-zone span { color: #3b82f6; font-weight: 600; }
  .file-pill { display: inline-flex; align-items: center; gap: 8px; background: #f0fdf4; border: 1px solid #86efac; border-radius: 20px; padding: 4px 12px; font-size: 13px; color: #166534; margin-top: 10px; }
  .metrics { display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin-bottom: 12px; }
  .metric { background: #f8fafc; border-radius: 8px; padding: 14px 16px; border: 1px solid #e2e8f0; }
  .metric-label { font-size: 11px; color: #94a3b8; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 6px; }
  .metric-value { font-size: 26px; font-weight: 700; color: #1e293b; }
  .metric-value.warn { color: #d97706; }
  .metric-value.danger { color: #dc2626; }
  .anchor-info { font-size: 12px; color: #475569; background: #f0f9ff; border: 1px solid #bae6fd; border-radius: 6px; padding: 8px 12px; }
  .form-row { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; margin-bottom: 14px; }
  .form-row-3 { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 14px; margin-bottom: 14px; }
  .field label { display: block; font-size: 12px; color: #64748b; font-weight: 600; margin-bottom: 5px; }
  .field input { width: 100%; padding: 9px 12px; border: 1px solid #cbd5e1; border-radius: 6px; font-size: 14px; color: #1e293b; background: #fff; }
  .field input:focus { outline: none; border-color: #3b82f6; box-shadow: 0 0 0 3px rgba(59,130,246,0.1); }
  .field .hint { font-size: 11px; color: #94a3b8; margin-top: 4px; }
  .btn-row { display: flex; gap: 10px; margin-top: 4px; }
  .btn { padding: 10px 22px; border-radius: 7px; font-size: 14px; font-weight: 600; cursor: pointer; border: none; transition: opacity 0.15s, transform 0.1s; }
  .btn:active { transform: scale(0.98); }
  .btn:disabled { opacity: 0.4; cursor: not-allowed; }
  .btn-preview { background: #1e293b; color: #fff; }
  .btn-send { background: #16a34a; color: #fff; }
  .btn-preview:hover:not(:disabled) { background: #0f172a; }
  .btn-send:hover:not(:disabled) { background: #15803d; }
  .status { font-size: 13px; margin-top: 10px; padding: 6px 10px; border-radius: 5px; }
  .status.info { background: #eff6ff; color: #1d4ed8; }
  .status.ok { background: #f0fdf4; color: #166534; }
  .status.err { background: #fef2f2; color: #991b1b; }
  .preview-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 12px; }
  .btn-copy { padding: 6px 14px; border: 1px solid #cbd5e1; border-radius: 6px; background: #fff; font-size: 12px; cursor: pointer; font-weight: 600; color: #475569; }
  .btn-copy:hover { background: #f8fafc; }
  #preview-box { font-family: 'Courier New', monospace; font-size: 12px; line-height: 1.65; background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 16px; white-space: pre; overflow-x: auto; max-height: 480px; overflow-y: auto; color: #1e293b; }
  .spinner { display: inline-block; width: 14px; height: 14px; border: 2px solid rgba(255,255,255,0.4); border-top-color: #fff; border-radius: 50%; animation: spin 0.6s linear infinite; vertical-align: middle; margin-right: 6px; }
  @keyframes spin { to { transform: rotate(360deg); } }
  .hidden { display: none !important; }
  @media (max-width: 600px) { .metrics { grid-template-columns: repeat(2,1fr); } .form-row,.form-row-3 { grid-template-columns: 1fr; } }
</style>
</head>
<body>
<div class="topbar"><div class="dot"></div><h1>Device Expiry Report — Email Sender</h1></div>
<div class="layout">

  <div class="card">
    <p class="card-title">1 · Upload Excel file</p>
    <div class="upload-zone" id="drop-zone" onclick="document.getElementById('xlsxFile').click()"
         ondragover="ev(event,'drag')" ondragleave="ev(event,'')" ondrop="drop(event)">
      <input type="file" id="xlsxFile" accept=".xlsx,.xls" onchange="fileChosen(this.files[0])">
      <div class="upload-icon">📂</div>
      <p><span>Click to browse</span> or drag &amp; drop</p>
      <p style="font-size:12px;color:#94a3b8;margin-top:4px">Must contain: Company · Plate Number · Installation Date</p>
    </div>
    <div id="file-pill" class="hidden"><div class="file-pill" id="pill-text"></div></div>
  </div>

  <div class="card hidden" id="metrics-card">
    <p class="card-title">Report snapshot</p>
    <div class="metrics">
      <div class="metric"><div class="metric-label">Total devices</div><div class="metric-value" id="m-total">—</div></div>
      <div class="metric"><div class="metric-label">Upcoming</div><div class="metric-value warn" id="m-up">—</div></div>
      <div class="metric"><div class="metric-label">Last 30 days</div><div class="metric-value danger" id="m-b1">—</div></div>
      <div class="metric"><div class="metric-label">Last 60 days</div><div class="metric-value" id="m-b2">—</div></div>
    </div>
    <div class="anchor-info" id="anchor-info"></div>
  </div>

  <div class="card">
    <p class="card-title">2 · Configure &amp; send</p>
    <div class="form-row-3">
      <div class="field">
        <label>Anchor date (reference)</label>
        <input type="text" id="anchor-input" placeholder="01-04-2026">
        <div class="hint">Must be 1st or 16th of month · dd-mm-yyyy</div>
      </div>
      <div class="field">
        <label>To</label>
        <input type="email" id="to-addr" value="{{ default_to }}" placeholder="recipient@example.com">
      </div>
      <div class="field">
        <label>CC (optional)</label>
        <input type="email" id="cc-addr" value="{{ default_cc }}" placeholder="cc@example.com">
      </div>
    </div>
    <div class="form-row">
      <div class="field">
        <label>From (sender)</label>
        <input type="email" id="from-addr" value="{{ gmail_user }}" readonly style="background:#f8fafc;color:#64748b">
      </div>
      <div class="field">
        <label>Report date (for subject line)</label>
        <input type="text" id="report-date" placeholder="01-04-2026">
      </div>
    </div>
    <div class="btn-row">
      <button class="btn btn-preview" id="btn-preview" onclick="doPreview()" disabled>Preview email</button>
      <button class="btn btn-send" id="btn-send" onclick="doSend()" disabled>Send email</button>
    </div>
    <div class="status hidden" id="status-bar"></div>
  </div>

  <div class="card hidden" id="preview-card">
    <div class="preview-header">
      <p class="card-title" style="margin:0">3 · Email preview</p>
      <button class="btn-copy" onclick="copyPreview()">📋 Copy</button>
    </div>
    <div id="preview-box"></div>
  </div>
</div>

<script>
let chosenFile = null;

// Default anchor = today formatted as dd-mm-yyyy
const td = new Date();
const dd = String(td.getDate()).padStart(2,'0');
const mm = String(td.getMonth()+1).padStart(2,'0');
const yyyy = td.getFullYear();
const todayStr = `${dd}-${mm}-${yyyy}`;
document.getElementById('anchor-input').value = todayStr;
document.getElementById('report-date').value = todayStr;

function ev(e, cls) { e.preventDefault(); document.getElementById('drop-zone').className = 'upload-zone'+(cls?' '+cls:''); }
function drop(e) { e.preventDefault(); document.getElementById('drop-zone').className='upload-zone'; if(e.dataTransfer.files[0]) fileChosen(e.dataTransfer.files[0]); }
function fileChosen(f) {
  if (!f) return;
  chosenFile = f;
  document.getElementById('pill-text').textContent = '✓ '+f.name+'  ('+(f.size/1024).toFixed(0)+' KB)';
  document.getElementById('file-pill').classList.remove('hidden');
  document.getElementById('btn-preview').disabled = false;
  document.getElementById('btn-send').disabled = false;
  document.getElementById('metrics-card').classList.add('hidden');
  document.getElementById('preview-card').classList.add('hidden');
  setStatus('','');
}
function setStatus(msg, type) {
  const el = document.getElementById('status-bar');
  if (!msg) { el.classList.add('hidden'); return; }
  el.textContent = msg; el.className = 'status '+type; el.classList.remove('hidden');
}
function buildFormData() {
  const fd = new FormData();
  fd.append('excel', chosenFile);
  fd.append('anchor', document.getElementById('anchor-input').value.trim());
  fd.append('report_date', document.getElementById('report-date').value.trim());
  fd.append('to_email', document.getElementById('to-addr').value);
  fd.append('cc_email', document.getElementById('cc-addr').value);
  return fd;
}
function setBtns(loading) {
  document.getElementById('btn-preview').disabled = loading;
  document.getElementById('btn-send').disabled = loading;
}
async function doPreview() {
  if (!chosenFile) return;
  const anchor = document.getElementById('anchor-input').value.trim();
  if (!anchor) { setStatus('Please enter an anchor date (e.g. 01-04-2026)', 'err'); return; }
  setBtns(true); setStatus('Generating preview…','info');
  try {
    const res = await fetch('/preview', {method:'POST', body:buildFormData()});
    const data = await res.json();
    if (data.error) { setStatus('Error: '+data.error,'err'); setBtns(false); return; }
    document.getElementById('m-total').textContent = data.stats.total;
    document.getElementById('m-up').textContent = data.stats.upcoming;
    document.getElementById('m-b1').textContent = data.stats.b1;
    document.getElementById('m-b2').textContent = data.stats.b2;
    document.getElementById('anchor-info').textContent =
      'Anchor: '+data.stats.anchor+'  |  Upcoming window: '+data.stats.upcoming_window+'  |  Last 30d: '+data.stats.b1_range;
    document.getElementById('metrics-card').classList.remove('hidden');
    document.getElementById('preview-box').textContent = data.preview;
    document.getElementById('preview-card').classList.remove('hidden');
    setStatus('Preview ready. Review below, then click Send.','ok');
  } catch(e) { setStatus('Network error: '+e.message,'err'); }
  setBtns(false);
}
async function doSend() {
  if (!chosenFile) return;
  const to = document.getElementById('to-addr').value.trim();
  if (!to) { setStatus('Please enter a recipient email address.','err'); return; }
  const anchor = document.getElementById('anchor-input').value.trim();
  if (!anchor) { setStatus('Please enter an anchor date.','err'); return; }
  if (!confirm(`Send email to ${to}?`)) return;
  setBtns(true);
  document.getElementById('btn-send').innerHTML = '<span class="spinner"></span>Sending…';
  setStatus('Sending email with attachment…','info');
  try {
    const res = await fetch('/send', {method:'POST', body:buildFormData()});
    const data = await res.json();
    if (data.error) { setStatus('Error: '+data.error,'err'); }
    else { setStatus('✓ '+data.message,'ok'); }
  } catch(e) { setStatus('Network error: '+e.message,'err'); }
  document.getElementById('btn-send').textContent = 'Send email';
  setBtns(false);
}
function copyPreview() {
  navigator.clipboard.writeText(document.getElementById('preview-box').textContent).then(() => {
    const b = document.querySelector('.btn-copy');
    b.textContent = '✓ Copied!';
    setTimeout(() => b.textContent = '📋 Copy', 1800);
  });
}
</script>
</body>
</html>"""

# ── routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template_string(HTML_PAGE,
        default_to=DEFAULT_TO, default_cc=DEFAULT_CC, gmail_user=FROM_EMAIL)

@app.route("/preview", methods=["POST"])
def preview():
    file = request.files.get("excel")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400
    anchor_str      = request.form.get("anchor", "").strip()
    report_date_str = request.form.get("report_date", "").strip()
    try:
        anchor = parse_anchor(anchor_str)
        if anchor.day not in (1, 16):
            return jsonify({"error": f"Anchor day must be 1st or 16th of the month. Got: {anchor.day}"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    try:
        df = load_dataframe(file)
        buckets = bucket_records(df, anchor)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    up_start = buckets["_up_start"]
    up_end   = buckets["_up_end"]
    b1_start = buckets["_b1_start"]
    b1_end   = buckets["_b1_end"]
    plain = build_email_body(buckets, report_date_str or anchor_str)
    stats = {
        "total":          len(df),
        "upcoming":       len(buckets["upcoming"]),
        "b1":             len(buckets["b1"]),
        "b2":             len(buckets["b2"]),
        "anchor":         anchor.strftime("%d-%m-%Y"),
        "upcoming_window": date_range_label(up_start, up_end),
        "b1_range":        date_range_label(b1_start, b1_end),
    }
    return jsonify({"preview": plain, "stats": stats})

@app.route("/send", methods=["POST"])
def send():
    file = request.files.get("excel")
    if not file:
        return jsonify({"error": "No file uploaded"}), 400
    to_addr         = request.form.get("to_email", DEFAULT_TO).strip()
    cc_addr         = request.form.get("cc_email",  DEFAULT_CC).strip()
    anchor_str      = request.form.get("anchor", "").strip()
    report_date_str = request.form.get("report_date", "").strip()
    file_bytes = file.read()
    filename   = file.filename
    try:
        anchor = parse_anchor(anchor_str)
        if anchor.day not in (1, 16):
            return jsonify({"error": f"Anchor day must be 1st or 16th. Got: {anchor.day}"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    try:
        df      = load_dataframe(io.BytesIO(file_bytes))
        buckets = bucket_records(df, anchor)
    except Exception as e:
        return jsonify({"error": f"Parse error: {e}"}), 500

    label   = report_date_str or anchor_str
    plain   = build_email_body(buckets, label)
    html    = build_html_body(buckets, label)
    subject = f"Device Expiry Report - {label}"

    try:
        result_excel_bytes = build_result_excel(buckets, label)
        result_filename = f"Expiry_Report_{label.replace('-','_')}.xlsx"
    except Exception as e:
        return jsonify({"error": f"Result Excel build error: {e}"}), 500

    attachments = [
        {"filename": filename,        "content": file_bytes},
        {"filename": result_filename, "content": result_excel_bytes},
    ]
    try:
        send_email(to_addr, cc_addr, subject, plain, html, attachments)
    except Exception as e:
        return jsonify({"error": f"Email error: {e}"}), 500
    return jsonify({"success": True, "message": f"Email sent to {to_addr} with 2 attachments"})

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
